import openpyxl

from openpyxl import workbook, load_workbook

wb1 = load_workbook(r'C:\Users\koeni\Desktop\VSC\AmpLabels\Excel sheets for Mabellaker\Triple\162227-21_Savage_Amp_Patch_001.xlsx')
wb2 = load_workbook(r'C:\Users\koeni\Desktop\VSC\AmpLabels\Excel sheets for Mabellaker\Double\AESPA Amp patch-001.xlsx')

ws = wb1.active
ws2 = wb2.active 
ws3 = wb1["Table 2"]
ws4 = wb2["Table 2"]

print('SR Amp Cart')

for i in range(8,20):
    if (ws[f'R{i}'].value) == (ws2[f'R{i}'].value) and (ws[f'Q{i}'].value) == (ws2[f'Q{i}'].value): {}
    else: 
        print(ws[f'A{i}'].value, ws[f'Q{i}'].value, ws[f'R{i}'].value)
        #output.append(        (ws[f'A{i}'].value +  "," +  ws[f'B{i}'].value + "," +  ws[f'C{i}'].value)

print()

print('SL Amp Cart')

print()

for j in range(8,20):
    if (ws3[f'R{j}'].value) == (ws4[f'R{j}'].value) and (ws3[f'Q{j}'].value) == (ws4[f'Q{j}'].value): {}
    else: 
        print(ws3[f'A{j}'].value, ws3[f'Q{j}'].value, ws3[f'R{j}'].value)